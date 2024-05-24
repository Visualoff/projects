import os
import json
import re
import tkinter as tk

import requests
import fake_useragent
from bs4 import BeautifulSoup

import openpyxl
from googletrans import Translator
'''Приложение для извлечения слов с сайтов для их дальнейшего перевода
    и сохранения в Excel таблицу'''

# Создание или запись форм идентификации для requests в .json формате
def json_data(form= 'CodeWars', id_log= 'user_email', id_pass= 'user_password'):

    data = {
        form: {
            'form_sent': '1',
            id_log: 'login',
            id_pass: 'password'
        }
    }

    data_old = {}

    with open('data.json', 'r') as f:
        try:
            data_old = json.load(f)
        except json.JSONDecodeError:
            pass

    with open('data.json', 'w') as f:
        if data_old:
            data_old.update(data)
            json.dump(data_old, f, indent=4)
        else:
            json.dump(data, f, indent=4)

# Создание сессии и выгрузка текста с сайта
def pars_login_url(pars_link, link=None, login=None, password=None):

    s = requests.Session()

    header = {
        'user-agent': fake_useragent.UserAgent().random
    }

    # Загрузка форм идентификации с заменой значений ключей авторизации
    with open('data.json', 'r') as f:
        data = dict(json.load(f))
        for form in data.values():
            for id_keys in form.keys():
                if form[id_keys] == 'login':
                    form[id_keys] = login
                if form[id_keys] == 'password':
                    form[id_keys] = password

    # Вход на сайт
    if link:
        try:
            s.post(link, data=data, headers= header)
        except requests.exceptions.MissingSchema:
            return "Include http or https"
        except requests.exceptions.HTTPError:
            return "HTTP Error"
        except requests.exceptions.ReadTimeout:
            return "Time out"
        except requests.exceptions.ConnectionError:
            return "Connection error"
        except requests.exceptions.RequestException:
            return "Exception request"
    if pars_link:
        try:
            text_html = s.get(pars_link, headers= header).content
            soup = BeautifulSoup(text_html, 'html.parser')
        except requests.exceptions.MissingSchema:
            return "Include http or https"
        except requests.exceptions.HTTPError:
            return "HTTP Error"
        except requests.exceptions.ReadTimeout:
            return "Time out"
        except requests.exceptions.ConnectionError:
            return "Connection error"
        except requests.exceptions.RequestException:
            return "Exception request"

        # Получение всего текста со страницы включая вложенные параграфы используя рекурсию
        def get_all_text(soup_html):

            text = ""

            if isinstance(soup_html, str):
                return soup_html

            for child in soup_html.children:
                if child.name is not None:
                    text += get_all_text(child)

            if soup_html.name is not None and soup_html.get_text(strip=True):
                text += ' ' + soup_html.get_text(strip=True)
            return text

        words = get_all_text(soup)
        words = re.findall(r'\b\w+\b', words.lower())

        # Фильтрация текста: Только латиница, без цифр и символов.
        re_words = []
        for word in words:
            if not re.search(r'\d', word) and 1 < len(word) <= 13:
                re_words.append(re.sub(r'[^\w\s]', '', word))
        latin_regex = re.compile(r'^[a-zA-Z]+$')
        filtred_words = list(set([word for word in re_words if latin_regex.match(word)]))
        return filtred_words
#Перевод и запись слов в .xlsx формат
def translate_and_add_to_excel(words):
    try:
        wb = openpyxl.load_workbook('words.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    ws = wb.active

    # Фильтрация от уже записанных слов.
    filtered_list = []
    words_in_excel = [cell.value for cell in ws['A']]
    for word in words:
        if word not in words_in_excel:
            filtered_list.append(word)

    # Перевод слов и запись их в словарь
    # Цикл нужен для отправки слов небольшими списками что бы googletrans не перегружался
    translated_list = []
    copy_filtered_list = filtered_list.copy()
    counter = len(filtered_list)
    while counter != 0:

        translator = Translator()
        filtered_words = copy_filtered_list[0:10]

        for word in filtered_words:
            translated_word = translator.translate(word, src='en', dest='ru').text
            translated_list.append(translated_word)
            copy_filtered_list.remove(word)
            counter -= 1
    translated_dict = dict(zip(filtered_list, translated_list))

    # Добавление переведенных слов в 'words.xlsx' если они переведены
    cyrillic_regex = re.compile(r'^[А-Яа-яЁё]+$')
    for key, value in translated_dict.items():
        if cyrillic_regex.match(value):
            print(f'Запись слова: {key}: {value}')
            ws.append([key, value])

    # Сохранение обьекта в words.xlsx
    try:
        wb.save('words.xlsx')
    except PermissionError:
        return 'Please close words.xlsx'

# Что то наподобие GUI для этого скрипта
def window():
    window = tk.Tk()
    window.title("LinkTranslator")
    window.geometry('400x90+500+500')
    window.resizable(width=False, height=False)
    window['bg'] = '#1f1f1f'
    window.iconbitmap('icon.ico')

    frame_button = tk.Frame(window, background= '#1f1f1f')
    frame_button.pack()

    image = tk.PhotoImage(file='logo.png')

    label_image = tk.Label(frame_button, bg='#1f1f1f', font= 'Arial 15', image=image)
    label_image.pack(side=tk.TOP)
    def any_page_win():
        window = tk.Toplevel()
        window.title("LinkTranslator - any")
        window.geometry('650x300+500+500')
        window.resizable(width=False, height=False)
        window['bg'] = '#1f1f1f'
        window.iconbitmap('icon.ico')

        frame_input = tk.Frame(window, background= '#1f1f1f')
        frame_input.pack()

        label_1 = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc')
        label_1.pack()

        label_ep = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Extracted page')
        label_ep.pack()

        entry_pars_link = tk.Entry(frame_input, width=100, bg= '#9d9d9d')
        entry_pars_link.pack()

        label_link = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Login page')
        label_link.pack()

        entry_link = tk.Entry(frame_input, width=100, bg= '#9d9d9d')
        entry_link.pack()

        label_login = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Login')
        label_login.pack()

        entry_login = tk.Entry(frame_input, width=100, bg= '#9d9d9d')
        entry_login.pack()

        label_pass = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Password')
        label_pass.pack()

        entry_password = tk.Entry(frame_input, width=100, bg= '#9d9d9d')
        entry_password.pack()

        label_2 = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc')
        label_2.pack()

        def command_ok():
            pars_link = entry_pars_link.get()
            link = entry_link.get()
            login = entry_login.get()
            password = entry_password.get()
            list_post = [pars_link, link, login, password]
            words = pars_login_url(*list_post)
            if type(words) == type([]) and words != []:
                translate_and_add_to_excel(words)
                label['fg'] = 'green'
                label['text'] = 'Успешно!'
            else:
                label['fg'] = 'red'
                label['text'] = words
            entry_pars_link.delete(0, tk.END)

        button_ok = tk.Button(frame_input, text="ОК", width=40, bg= '#9d9d9d', command= command_ok)
        button_ok.pack()

        label = tk.Label(window, bg='#1f1f1f', font= 'Arial 15')
        label.pack()

    def codewars_win():
        window = tk.Toplevel()
        window.title("LinkTranslator - CodeWars")
        window.geometry('650x150+500+500')
        window.resizable(width=False, height=False)
        window['bg'] = '#1f1f1f'
        window.iconbitmap('icon.ico')

        frame_input = tk.Frame(window, background= '#1f1f1f')
        frame_input.pack()

        label_1 = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc')
        label_1.pack()

        # Вход по данным из log.txt если он есть и его удаление, иначе создание log.txt с вашими данными
        if os.path.exists('log.txt'):
            label_ep = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Extracted page')
            label_ep.pack()

            entry_pars_link = tk.Entry(frame_input, width=100, bg= '#9d9d9d')
            entry_pars_link.pack()

            label_2 = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc')
            label_2.pack()

            def command_ok():
                pars_link = entry_pars_link.get()
                link = 'https://www.codewars.com/users/sign_in'

                with open('log.txt', 'r') as file:
                    loaded_log = file.readlines()
                    loaded_log = [s.strip() for s in loaded_log]

                words = pars_login_url(pars_link, link, *loaded_log)

                if type(words) == type([]) and words != []:
                    translate_and_add_to_excel(words)
                    label['fg'] = 'green'
                    label['text'] = 'Успешно!'
                else:
                    label['fg'] = 'red'
                    label['text'] = words

                entry_pars_link.delete(0, tk.END)

            def command_del_log():
                os.remove('log.txt')
                window.destroy()
                codewars_win()

            button_del_log = tk.Button(frame_input, text="Delete log", width=40, bg= '#9d9d9d', command= command_del_log)
            button_del_log.pack(side= tk.LEFT)

            label = tk.Label(frame_input, bg='#1f1f1f', font= 'Arial 15', text= '    ')
            label.pack(side= tk.LEFT)

            button_ok = tk.Button(frame_input, text="ОК", width=40, bg= '#9d9d9d', command= command_ok)
            button_ok.pack(side= tk.LEFT)

            label = tk.Label(window, bg='#1f1f1f', font= 'Arial 15')
            label.pack()
        else:
            window.geometry('350x170+500+500')

            label_log = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Login')
            label_log.pack()

            entry_log = tk.Entry(frame_input, width=50, bg= '#9d9d9d')
            entry_log.pack()

            label_pass = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Password')
            label_pass.pack()

            entry_pass = tk.Entry(frame_input, width=50, bg= '#9d9d9d')
            entry_pass.pack()

            label_2 = tk.Label(frame_input, bg='#1f1f1f', fg='#cccccc')
            label_2.pack()

            def command_ok():
                with open('log.txt', 'w') as file:
                    file.write(entry_log.get() + '\n')
                    file.write(entry_pass.get() + '\n')
                window.destroy()
                codewars_win()

            button_ok = tk.Button(frame_input, text="ОК", width=40, bg= '#9d9d9d', command= command_ok)
            button_ok.pack()

    def json_win():
        win = tk.Toplevel()
        win.title(" ")
        win.geometry('180x220+500+300')
        win.resizable(width=False, height=False)
        win['bg'] = '#1f1f1f'
        win.iconbitmap('icon.ico')

        label_1 = tk.Label(win, bg='#1f1f1f', fg='#cccccc')
        label_1.pack()

        label_name = tk.Label(win, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='Name Form')
        label_name.pack()

        entry_name = tk.Entry(win, width=20, bg= '#9d9d9d')
        entry_name.pack()

        label_user = tk.Label(win, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='USER_id')
        label_user.pack()

        entry_user = tk.Entry(win, width=20, bg= '#9d9d9d')
        entry_user.pack()

        label_pass = tk.Label(win, bg='#1f1f1f', fg='#cccccc', font='Arial 12', text='PASS_id')
        label_pass.pack()

        entry_pass = tk.Entry(win, width=20, bg= '#9d9d9d')
        entry_pass.pack()

        label_2 = tk.Label(win, bg='#1f1f1f', font= 'Arial 15')
        label_2.pack()

        def json_save():
            form = entry_name.get()
            user_id = entry_user.get()
            pass_id = entry_pass.get()
            json_data(form, user_id, pass_id)
            win.destroy()

        button_add = tk.Button(win, text="Save", width=10, bg= '#9d9d9d', command=json_save)
        button_add.pack()

    button_any = tk.Button(frame_button, text="Any", width=20, bg= '#9d9d9d', command=any_page_win)
    button_any.pack(side=tk.LEFT)

    button_cd = tk.Button(frame_button, text="CodeWars", width=20, bg= '#9d9d9d', command= codewars_win)
    button_cd.pack(side=tk.LEFT)

    button_json = tk.Button(frame_button, text="JSON", width=20, bg= '#9d9d9d', command=json_win)
    button_json.pack(side=tk.LEFT)

    window.mainloop()

if __name__ == '__main__':
    if not os.path.exists('data.json'):
        with open('data.json', 'w')as f:
            f.close()
        json_data()
    window()
